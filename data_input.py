import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
import os

SUBJECTS = [
    "Physics 2",
    "Circuit Theory",
    "Analysis 2",
    "Algorithms and DS",
    # "Physics 1",
    # "LAG",
    # "Prog Tech",
    # "Chemistry",
]

RANGES = [0, 7, 20]
# RANGES = [0,1,7,20]
# RANGES = [0,1,7,15,30]

FILE_PATH = "g:/My Drive/study_schedule.xlsx"
LOCAL_FOLDER = "Python/study_schedule"


def subject_selection(subjects):
    """returns subject chosen by the user"""
    counter = 0
    for subject in subjects:
        counter += 1
        print(f"\n{counter}. {subject}", end="")

    found = False
    while not found:
        subject_choice = input("\nSELECT A SUBJECT: ").strip()
        if subject_choice.isdigit() and (1 <= int(subject_choice) <= len(subjects)):
            subject_choice = int(subject_choice) - 1
            found = True
        else:
            print("\nInvalid input!")

    active_subject = subjects[subject_choice]
    return active_subject


def user_interaction(subject):
    """adds topic(s); updates review; saves data to csv"""

    with open(f"{LOCAL_FOLDER}/{subject}.csv", "r+") as csv_file:
        first_line = csv_file.readline()
        # creating a dictionary containing all data in the csv
        # later the csv is overwritten with the updated dictionary
        topics = {
            line.strip().split(",")[0]: (
                line.strip().split(",")[1:]
                if line.strip().split(",")[1:] != [""]
                else []
            )
            for line in csv_file
        }

        while True:
            print(f"\nSubject: {subject}")
            print("1. Add topic(s)\n2. Update review")
            action_choice = input("SELECT AN ACTION: ").strip()

            if action_choice == "1":  # case 1
                topics_input = input("TOPIC NAME(S): ").strip()
                topics_list = [
                    topic.strip().lower() for topic in topics_input.split(",")
                ]

                for topic in topics_list:
                    if topic not in topics:
                        topics[topic] = []
                        initial_schedule(topic, topics)
                        print(f"'{topic}' was added")
                    else:
                        print(f"\n'{topic}' is already present!")

            elif action_choice == "2":  # case 2
                if topics:  # if dict not empty
                    print("\nAvailable Topics:")
                    counter = 0
                    key_finder = []
                    for topic in topics:
                        counter += 1
                        print(f"{counter}. {topic}")
                        key_finder.append((counter, topic))

                    update_topic = input("SELECT TOPIC: ").strip().casefold()
                    if update_topic in topics:
                        retrospective_schedule(update_topic, topics)

                    elif update_topic.isdigit() and (
                        1 <= int(update_topic) <= (len(key_finder))
                    ):
                        for count, key in key_finder:
                            if count == int(update_topic):
                                retrospective_schedule(key, topics)
                    else:
                        print("\nInvalid input!")
                else:
                    print("\nThere are no topics yet!")

            else:  # case 3 (exit) and 4 (invalid input)
                if action_choice == "":
                    placement = len("".join(first_line))
                    csv_file.seek(placement)  # placing cursor at second line
                    for topic, dates in topics.items():
                        csv_file.write(f"{topic},{','.join(dates)}\n")
                    csv_file.truncate()
                    return topics
                print("\nInvalid input!")


def actual_ranges(ranges):
    """computes ranges as delta from day 1"""

    actual_ranges = [sum(ranges[: ranges.index(i) + 1]) for i in ranges]
    return actual_ranges


def initial_schedule(topic, topics):
    """creates schedule based on user-selected start date"""

    found = False
    while not found:
        date_str = input("DATE (t/yyyy-mm-dd): ").strip().casefold()
        if date_str == "t":
            start_date = datetime.date.today()
            found = True
        else:
            try:
                datetime.datetime.strptime(date_str, "%Y-%m-%d")
                start_date = datetime.date(
                    int(date_str[:4]), int(date_str[5:7]), int(date_str[8:])
                )
                found = True
            except ValueError:
                print(
                    "\nInvalid input! Please enter 't' for today or a valid date (YYYY-MM-DD)."
                )

    for days in actual_ranges(RANGES):
        topics[topic].append(str(start_date + datetime.timedelta(days=days)))

    # """creates schedule based on day the topic is added"""

    # for range in actual_ranges(RANGES):
    #    topics[topic].append(str(datetime.date.today() +
    #                             datetime.timedelta(days=range)))


def retrospective_schedule(topic, topics):
    """creates schedule based on review date and rating"""

    topics[topic] = topics[topic][: len(RANGES)]  # trim extra dates
    counter = "".join(topics[topic]).count(";")

    if 0 <= counter < len(RANGES):  # if not rated yet
        found = False
        while not found:
            date_str = (
                input(f"REVIEW n{counter + 1} (y/t/yyyy-mm-dd): ").strip().casefold()
            )
            if date_str == "t":
                date = datetime.date.today()
                found = True
            elif date_str == "y":
                date = datetime.date.today() - datetime.timedelta(days=1)
                found = True
            else:
                try:
                    datetime.datetime.strptime(date_str, "%Y-%m-%d")
                    date = datetime.date(
                        int(date_str[:4]), int(date_str[5:7]), int(date_str[8:])
                    )
                    found = True
                except ValueError:
                    print("\nInvalid input!")

        found = False
        while not found:
            try:
                rating = int(input("RATING 1-4: ").strip())
                if 1 <= rating <= 4:
                    found = True
                else:
                    print("\nInvalid Input!")
            except ValueError:
                print("\nInvalid Input!")

        topics[topic][counter] = f"{date};{rating}"

        new_ranges = actual_ranges(rating_formula(rating)[counter + 1 :])
        for range in new_ranges:
            counter += 1
            while len(topics[topic]) <= counter:  # in case there are too few dates
                topics[topic].append("")
            topics[topic][counter] = str(date + datetime.timedelta(days=range))

    else:
        print("\nMax n of reviews reached!")


def rating_formula(rating):
    """returns rated ranges"""

    if rating == 1:
        modifier = 0.5
    elif rating == 2:
        modifier = 0.75
    elif rating == 3:
        modifier = 1  # default
    else:
        modifier = 1.5

    rated_ranges = [0] + [max(1, round(range * modifier)) for range in RANGES[1:]]
    return rated_ranges


def upload_to_excel(subject, topics):
    """uploads csv files to excel"""

    if os.path.exists(FILE_PATH):
        workbook = openpyxl.load_workbook(FILE_PATH)
    else:
        workbook = openpyxl.Workbook()
        if workbook.active:
            workbook.remove(workbook.active)

    if subject in workbook.sheetnames:
        del workbook[subject]
    sheet = workbook.create_sheet(title=subject)

    # sheet.append(["Topic"] + list(topics.keys()))
    sheet.append(["Topic"] + [f"Review {i + 1}" for i in range(len(RANGES))])

    color_map = {"1": "e06666", "2": "f79646", "3": "92d050", "4": "00b050"}

    # for row_idx, dates in enumerate(zip(*topics.values()), start=2):
    # sheet.cell(row=row_idx, column=1, value=f"Review {row_idx-1}")
    for row_idx, (topic, dates) in enumerate(topics.items(), start=2):
        sheet.cell(row=row_idx, column=1, value=topic)

        for col_idx, entry in enumerate(dates, start=2):
            if ";" in entry:
                date, rating = entry.split(";")
                cell = sheet.cell(row=row_idx, column=col_idx, value=date)
                if rating in color_map:
                    cell.fill = PatternFill(
                        start_color=color_map[rating], fill_type="solid"
                    )
            else:
                sheet.cell(row=row_idx, column=col_idx, value=entry)

    # sheet.auto_filter.ref = f"B1:{sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate}"

    for col in sheet.columns:  # auto fit
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[col_letter].width = max_length + 2  # padding

    bold_font = Font(bold=True)  # conditional formatting
    today_formula = 'TEXT(TODAY(), "YYYY-MM-DD")'
    rule = FormulaRule(
        formula=[f'TEXT(A1, "YYYY-MM-DD")={today_formula}'], font=bold_font
    )

    sheet.conditional_formatting.add(
        f"A1:{sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate}",
        rule,
    )
    workbook.save(FILE_PATH)


def summary_scheduler():
    """adds summary sheet to excel"""

    today = datetime.date.today()
    tomorrow = today + datetime.timedelta(days=1)
    day_after = today + datetime.timedelta(days=2)

    overdue = []
    due_today = []
    due_tomorrow = []
    due_day_after = []

    for subject in SUBJECTS:
        try:
            with open(f"{LOCAL_FOLDER}/{subject}.csv") as csv_file:
                csv_file.readline()
                for line in csv_file:
                    line = line.strip().split(",")
                    topic = line[0]
                    dates = line[1:]

                    for date in dates:
                        if ";" not in date:
                            if date[:10] < str(today):
                                days_overdue = (
                                    today
                                    - datetime.datetime.strptime(
                                        date[:10], "%Y-%m-%d"
                                    ).date()
                                ).days
                                if days_overdue < 60:
                                    overdue.append((topic, subject))
                                    break
                            if str(today) in date:
                                due_today.append((topic, subject))
                                break
                            if str(tomorrow) in date:
                                due_tomorrow.append((topic, subject))
                                break
                            if str(day_after) in date:
                                due_day_after.append((topic, subject))
                                break
        except FileNotFoundError:
            continue

    workbook = openpyxl.load_workbook(FILE_PATH)

    if "Summary" in workbook.sheetnames:
        del workbook["Summary"]
    sheet = workbook.create_sheet(title="Summary")

    sheet.append(["", "Topic", "Subject"])

    def add_rows(category, data):
        if data:
            sheet.append([category, "", ""])
            for subject, topic in data:
                sheet.append(["", subject, topic])

    add_rows("Overdue", overdue)
    add_rows(f"Due {today}", due_today)
    add_rows(f"Due {tomorrow}", due_tomorrow)
    add_rows(f"Due {day_after}", due_day_after)

    # sheet.auto_filter.ref = f"B1:C{sheet.max_row}"

    for col in sheet.columns:  # auto fit
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 1  # padding

    workbook.save(FILE_PATH)


def main():
    subject = subject_selection(SUBJECTS)

    try:
        topics = user_interaction(subject)
    except FileNotFoundError:  # if file does not exists yet
        f = open(f"{LOCAL_FOLDER}/{subject}.csv", "w")
        f.write("topic,")
        reviews = []
        for i in range(len(RANGES)):
            reviews.append(f"review {i + 1}")
        f.write(f"{','.join(reviews)}\n")
        # write in the first line because dictionary starts from the second
        f.close()
        topics = user_interaction(subject)

    upload_to_excel(subject, topics)
    summary_scheduler()


if __name__ == "__main__":
    main()
